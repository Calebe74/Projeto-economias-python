import openpyxl

# demorei pra achar o erro...
# essa funçao carrega a planilha e coloca os dados novos na linha a baixo dos dados que ja tem la
def estatistica():
    wb = openpyxl.load_workbook("dados.xlsx")
    sheet = wb["Resultado da CL"]

    primeira_linha = 2
    ultima_linha = sheet.max_row

    custo_total_total = 0
    lucro_total = 0
    preco_venda_total = 0

    for linha in range(primeira_linha, ultima_linha + 1):
        custo_total = sheet.cell(row=linha, column=7).value
        lucro = sheet.cell(row=linha, column=8).value
        preco_venda = sheet.cell(row=linha, column=3).value

        custo_total_total += custo_total if custo_total is not None else 0
        lucro_total += lucro if lucro is not None else 0
        preco_venda_total += preco_venda if preco_venda is not None else 0

    margem_lucro_total = (lucro_total / preco_venda_total * 100) if preco_venda_total > 0 else 0

    custo_total = round(custo_total_total, 2)
    lucro_liquido_total = round(lucro_total, 2)
    margem_de_lucro_total = round(margem_lucro_total, 2)

    return [custo_total, lucro_liquido_total, margem_de_lucro_total]


# essa pega os dados da planilha pra deixar na interfacie
def obter_dados_excel(nome_arquivo):
    wb = openpyxl.load_workbook(nome_arquivo)
    sheet = wb.active
    dados = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        dados.append(row)

    return dados


# ----------------------------------------

# salva o produto q o usuario colocar na interfacie
def salvar_produto(nome_produto,preco_compra,preco_venda,quantidade,custos_adicionais,custo_frete):
    
    # onde salva os dados q o usuario colocar
    nome_produto = nome_produto
    preco_compra = float(preco_compra)
    preco_venda = float(preco_venda)
    quantidade = int(quantidade)
    custos_adicionais = float(custos_adicionais)
    custo_frete = float(custo_frete)

    # calcula o lucro
    custo_total = (preco_compra + custos_adicionais + custo_frete) * quantidade
    lucro = (preco_venda - preco_compra - custos_adicionais - custo_frete) * quantidade
    margem_lucro = lucro / (preco_venda * quantidade) * 100

    # aqui se existir uma planilha com o nome q eu colocar o programa abre ela 
    # e se nao tiver, cria uma nova com os parametros que eu coloquei
    try:
        wb = openpyxl.load_workbook("dados.xlsx")
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        wb.active.title = "Resultado da Calculadora de Lucro"
        wb.active.append(["Produto", "Preço de compra", "Preço de venda", "Quantidade", "Custos adicionais", "Custo médio do frete", "Custo total", "Lucro líquido", "Margem de lucro"])

    sheet = wb["Resultado da CL"]
    last_row = sheet.max_row + 1

    sheet.cell(row=last_row, column=1).value = nome_produto
    sheet.cell(row=last_row, column=2).value = preco_compra
    sheet.cell(row=last_row, column=3).value = preco_venda
    sheet.cell(row=last_row, column=4).value = quantidade
    sheet.cell(row=last_row, column=5).value = custos_adicionais
    sheet.cell(row=last_row, column=6).value = custo_frete
    sheet.cell(row=last_row, column=7).value = custo_total
    sheet.cell(row=last_row, column=8).value = lucro
    sheet.cell(row=last_row, column=9).value = margem_lucro

    # salva a planilha com os dados novos
    wb.save("dados.xlsx")


# funçao para deletar a linha do produto selecionado
def deletar_linha_por_nome(nome_produto, nome_planilha):
    wb = openpyxl.load_workbook(nome_planilha)
    sheet = wb.active
    contador = 2

    for row in sheet.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
        if str(row[0]) == nome_produto:
            linha = contador
            sheet.delete_rows(linha)
            break

        # incrementando o contador
        contador +=1

    wb.save(nome_planilha)
